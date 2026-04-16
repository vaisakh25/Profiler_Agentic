"""
Layer 6 — Excel Profiling Engine (XLSX / XLS)

Steps:
  A — Open workbook and detect sheets; select the first (active) sheet.
  B — Header detection: heuristic on first 5 rows (same logic as CSV engine).
  C — Row count: exact count from sheet dimensions.
  D — Sampling: full read (MEMORY_SAFE) or read_only streaming (LAZY_SCAN / STREAM_ONLY).
  E — Build RawColumnData (pivot rows → columns).

Library used: openpyxl (for XLSX).  XLS (legacy .xls) is supported via xlrd
fallback if installed; otherwise raises NotImplementedError.

Design constraints:
  - Excel files > 100 MB are rare; warn if encountered but proceed.
  - Multi-sheet files: profiles the FIRST sheet only (consistent with the
    one-profile-per-file contract).  Sheet name is logged for context.
  - Merged cells are treated as if only the top-left cell holds the value.
  - Formula cells return their cached value (data_only=True).

Entry point:
  profile(path, strategy, intake=None) -> tuple[list[RawColumnData], int, bool]

Returns:
  (raw_columns, row_count, is_row_count_exact)
"""

from __future__ import annotations

import logging
import random
from pathlib import Path
from typing import Any, Optional

from file_profiler.config import settings
from file_profiler.intake.errors import CorruptFileError
from file_profiler.intake.validator import IntakeResult
from file_profiler.models.enums import SizeStrategy
from file_profiler.models.file_profile import RawColumnData
from file_profiler.observability.langsmith import compact_text_output, traceable

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Entry point\n# ---------------------------------------------------------------------------
@traceable(
    name="engine.excel.profile",
    run_type="chain",
    process_outputs=compact_text_output,
)
def profile(
    path: str | Path,
    strategy: SizeStrategy,
    intake: Optional[IntakeResult] = None,
) -> tuple[list[RawColumnData], int, bool]:
    """
    Profile an Excel file and return per-column raw data for the column profiler.

    Args:
        path:     Path to the Excel file (.xlsx or .xls).
        strategy: Size strategy selected by Layer 3.
        intake:   IntakeResult from Layer 1 (optional; kept for signature consistency).

    Returns:
        (raw_columns, row_count, is_row_count_exact)

    Raises:
        CorruptFileError — file cannot be opened or has no usable data.
        NotImplementedError — .xls file and xlrd is not installed.
    """
    path = Path(path).resolve()

    suffix = path.suffix.lower()
    if suffix == ".xls":
        return _profile_xls(path, strategy)

    # Default: XLSX via openpyxl
    return _profile_xlsx(path, strategy)


# ---------------------------------------------------------------------------
# XLSX profiling (openpyxl)
# ---------------------------------------------------------------------------

def _profile_xlsx(
    path: Path,
    strategy: SizeStrategy,
) -> tuple[list[RawColumnData], int, bool]:
    """Profile an XLSX file using openpyxl."""
    try:
        import openpyxl
    except ImportError as exc:
        raise NotImplementedError(
            "openpyxl is required for XLSX profiling. "
            "Install it with: pip install openpyxl"
        ) from exc

    use_read_only = strategy != SizeStrategy.MEMORY_SAFE

    try:
        wb = openpyxl.load_workbook(
            str(path),
            read_only=use_read_only,
            data_only=True,  # return cached formula values
        )
    except Exception as exc:
        raise CorruptFileError(
            f"Cannot open Excel file {path.name}: {exc}"
        ) from exc

    try:
        sheet_names = wb.sheetnames
        if not sheet_names:
            raise CorruptFileError(f"Excel file {path.name} has no sheets")

        ws = wb[sheet_names[0]]
        log.debug(
            "Excel: %s — sheet '%s' (%d of %d sheet(s))",
            path.name, sheet_names[0], 1, len(sheet_names),
        )

        if len(sheet_names) > 1:
            log.info(
                "Excel file %s has %d sheets; profiling first sheet '%s' only.",
                path.name, len(sheet_names), sheet_names[0],
            )

        # Read all rows as lists
        all_rows = _read_sheet_rows(ws, use_read_only)

        if not all_rows:
            log.warning("Excel sheet '%s' in %s has no data.", sheet_names[0], path.name)
            return [], 0, True

        # Step B — Header detection
        headers, has_header = _detect_headers(all_rows)

        # Step C — Row count
        data_rows = all_rows[1:] if has_header else all_rows
        row_count = len(data_rows)

        # Step D — Sampling
        sampled_rows = _sample_rows(data_rows, strategy)

        if not sampled_rows:
            log.warning("Excel engine: no rows sampled from %s", path.name)
            return [], row_count, True

        # Step E — Build RawColumnData
        raw_columns = _build_raw_columns(headers, sampled_rows, row_count)
        return raw_columns, row_count, True

    finally:
        wb.close()


def _read_sheet_rows(ws, read_only: bool) -> list[list[Any]]:
    """Read all rows from a worksheet, converting to a list of lists."""
    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        # Skip fully empty rows
        if all(cell is None for cell in row):
            continue
        rows.append(list(row))
    return rows


# ---------------------------------------------------------------------------
# XLS profiling (xlrd fallback)
# ---------------------------------------------------------------------------

def _profile_xls(
    path: Path,
    strategy: SizeStrategy,
) -> tuple[list[RawColumnData], int, bool]:
    """Profile a legacy XLS file using xlrd."""
    try:
        import xlrd
    except ImportError as exc:
        raise NotImplementedError(
            f"xlrd is required for legacy .xls profiling of '{path.name}'. "
            f"Install it with: pip install xlrd"
        ) from exc

    try:
        wb = xlrd.open_workbook(str(path))
    except Exception as exc:
        raise CorruptFileError(
            f"Cannot open XLS file {path.name}: {exc}"
        ) from exc

    if wb.nsheets == 0:
        raise CorruptFileError(f"XLS file {path.name} has no sheets")

    ws = wb.sheet_by_index(0)
    log.debug(
        "XLS: %s — sheet '%s' (%d of %d sheet(s))",
        path.name, ws.name, 1, wb.nsheets,
    )

    if wb.nsheets > 1:
        log.info(
            "XLS file %s has %d sheets; profiling first sheet '%s' only.",
            path.name, wb.nsheets, ws.name,
        )

    if ws.nrows == 0:
        return [], 0, True

    # Read all rows
    all_rows: list[list[Any]] = []
    for row_idx in range(ws.nrows):
        row = [ws.cell_value(row_idx, col_idx) for col_idx in range(ws.ncols)]
        if all(v == "" or v is None for v in row):
            continue
        all_rows.append(row)

    if not all_rows:
        return [], 0, True

    headers, has_header = _detect_headers(all_rows)
    data_rows = all_rows[1:] if has_header else all_rows
    row_count = len(data_rows)

    sampled_rows = _sample_rows(data_rows, strategy)
    if not sampled_rows:
        return [], row_count, True

    raw_columns = _build_raw_columns(headers, sampled_rows, row_count)
    return raw_columns, row_count, True


# ---------------------------------------------------------------------------
# Step B — Header Detection
# ---------------------------------------------------------------------------

def _detect_headers(rows: list[list[Any]]) -> tuple[list[str], bool]:
    """
    Determine whether the first row is a header or data.

    Heuristics (same as CSV engine):
    - All cells are non-numeric and no cell is blank → header present.
    - Otherwise → no header; generate column_1, column_2, ... names.
    """
    if not rows:
        return [], False

    first_row = rows[0]
    all_non_numeric = all(not _looks_numeric(cell) for cell in first_row)
    has_empty = any(_is_empty(cell) for cell in first_row)

    if all_non_numeric and not has_empty:
        headers = [_cell_to_header(cell) for cell in first_row]
        return _deduplicate_headers(headers), True

    n_cols = len(first_row)
    return [f"column_{i + 1}" for i in range(n_cols)], False


def _looks_numeric(value: Any) -> bool:
    """Check if a cell value looks numeric."""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    v = str(value).strip().replace(",", "")
    if not v:
        return False
    try:
        float(v)
        return True
    except (ValueError, TypeError):
        return False


def _is_empty(value: Any) -> bool:
    """Check if a cell is effectively empty."""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _cell_to_header(value: Any) -> str:
    """Convert a cell value to a header string."""
    if value is None:
        return ""
    return str(value).strip()


def _deduplicate_headers(headers: list[str]) -> list[str]:
    """Deduplicate header names by appending _2, _3, etc."""
    seen: dict[str, int] = {}
    result: list[str] = []
    for h in headers:
        if h not in seen:
            seen[h] = 0
            result.append(h)
        else:
            seen[h] += 1
            result.append(f"{h}_{seen[h] + 1}")
    return result


# ---------------------------------------------------------------------------
# Step D — Sampling
# ---------------------------------------------------------------------------

def _sample_rows(
    data_rows: list[list[Any]],
    strategy: SizeStrategy,
) -> list[list[Any]]:
    """Select rows based on the size strategy."""
    if strategy == SizeStrategy.MEMORY_SAFE:
        return data_rows

    if strategy == SizeStrategy.LAZY_SCAN:
        return _reservoir_sample(data_rows)

    return _skip_interval_sample(data_rows)


def _reservoir_sample(rows: list[list[Any]]) -> list[list[Any]]:
    """Vitter's Algorithm R — uniform random sample."""
    k = settings.SAMPLE_ROW_COUNT
    sample: list[list[Any]] = []
    rng = random.Random(42)

    for i, row in enumerate(rows):
        if i < k:
            sample.append(row)
        else:
            j = rng.randint(0, i)
            if j < k:
                sample[j] = row

    return sample


def _skip_interval_sample(rows: list[list[Any]]) -> list[list[Any]]:
    """Skip-interval sampling for STREAM_ONLY strategy."""
    interval = settings.STREAM_SKIP_INTERVAL
    return [row for i, row in enumerate(rows) if i % interval == 0]


# ---------------------------------------------------------------------------
# Step E — Build RawColumnData
# ---------------------------------------------------------------------------

def _build_raw_columns(
    headers: list[str],
    rows: list[list[Any]],
    total_count: int,
) -> list[RawColumnData]:
    """Pivot rows into columnar RawColumnData."""
    if not headers or not rows:
        return []

    n_cols = len(headers)
    columns: list[list[Optional[str]]] = [[] for _ in range(n_cols)]

    for row in rows:
        for col_idx in range(n_cols):
            value = row[col_idx] if col_idx < len(row) else None
            columns[col_idx].append(_cell_to_str(value))

    raw_cols: list[RawColumnData] = []
    for name, values in zip(headers, columns):
        null_count = sum(1 for v in values if v is None)
        raw_cols.append(RawColumnData(
            name=name,
            declared_type=None,
            values=values,
            total_count=total_count,
            null_count=null_count,
        ))

    return raw_cols


def _cell_to_str(value: Any) -> Optional[str]:
    """
    Convert an Excel cell value to Optional[str] for RawColumnData.

    - None → None
    - Empty string → None
    - bool → "true" / "false"
    - datetime → ISO format
    - Everything else → str()
    """
    if value is None:
        return None

    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None

    if isinstance(value, bool):
        return "true" if value else "false"

    # datetime.datetime, datetime.date, datetime.time
    if hasattr(value, "isoformat"):
        return value.isoformat()

    # Numeric: avoid trailing .0 for integers stored as float
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)

    return str(value)
