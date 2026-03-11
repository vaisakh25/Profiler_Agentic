"""
Tests for file_profiler/engines/excel_engine.py

Covers:
  - Basic XLSX profiling (header detection, row count, column values)
  - Numeric, string, boolean, date, and None cell types
  - Header detection heuristics (all-string vs numeric first row)
  - Generated headers when no header row
  - Multi-sheet files (first sheet profiled, others ignored)
  - Empty sheets
  - Duplicate headers
  - Sampling strategies
  - Float-to-int display (42.0 → "42")
  - Integration with main.profile_file
"""

from __future__ import annotations

import datetime
import os
import tempfile
from pathlib import Path

import pytest

from file_profiler.engines.excel_engine import (
    _cell_to_str,
    _detect_headers,
    _looks_numeric,
    profile,
)
from file_profiler.intake.errors import CorruptFileError
from file_profiler.models.enums import SizeStrategy


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _create_xlsx(rows: list[list], sheet_name: str = "Sheet1", extra_sheets: dict = None) -> Path:
    """Create a temp XLSX file with the given rows and return its path."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    if extra_sheets:
        for name, sheet_rows in extra_sheets.items():
            extra_ws = wb.create_sheet(title=name)
            for row in sheet_rows:
                extra_ws.append(row)
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    wb.close()
    return Path(path)


# ---------------------------------------------------------------------------
# Cell to String Conversion
# ---------------------------------------------------------------------------

class TestCellToStr:

    def test_none(self):
        assert _cell_to_str(None) is None

    def test_empty_string(self):
        assert _cell_to_str("") is None

    def test_whitespace_only(self):
        assert _cell_to_str("   ") is None

    def test_string(self):
        assert _cell_to_str("hello") == "hello"

    def test_string_stripped(self):
        assert _cell_to_str("  hello  ") == "hello"

    def test_int(self):
        assert _cell_to_str(42) == "42"

    def test_float_whole_number(self):
        assert _cell_to_str(42.0) == "42"

    def test_float_decimal(self):
        assert _cell_to_str(3.14) == "3.14"

    def test_bool_true(self):
        assert _cell_to_str(True) == "true"

    def test_bool_false(self):
        assert _cell_to_str(False) == "false"

    def test_datetime(self):
        dt = datetime.datetime(2024, 1, 15, 10, 30, 0)
        result = _cell_to_str(dt)
        assert "2024-01-15" in result

    def test_date(self):
        d = datetime.date(2024, 6, 15)
        assert _cell_to_str(d) == "2024-06-15"


# ---------------------------------------------------------------------------
# Header Detection
# ---------------------------------------------------------------------------

class TestHeaderDetection:

    def test_string_headers_detected(self):
        rows = [["Name", "Age", "City"], ["Alice", 30, "NYC"]]
        headers, has_header = _detect_headers(rows)
        assert has_header is True
        assert headers == ["Name", "Age", "City"]

    def test_numeric_first_row_no_header(self):
        rows = [[1, 2, 3], [4, 5, 6]]
        headers, has_header = _detect_headers(rows)
        assert has_header is False
        assert headers == ["column_1", "column_2", "column_3"]

    def test_empty_cell_in_first_row_no_header(self):
        rows = [["Name", None, "City"], ["Alice", 30, "NYC"]]
        headers, has_header = _detect_headers(rows)
        assert has_header is False

    def test_empty_rows(self):
        headers, has_header = _detect_headers([])
        assert headers == []
        assert has_header is False

    def test_duplicate_headers(self):
        rows = [["Name", "Name", "Age"], ["Alice", "Smith", 30]]
        headers, has_header = _detect_headers(rows)
        assert has_header is True
        assert headers == ["Name", "Name_2", "Age"]


class TestLooksNumeric:

    def test_int(self):
        assert _looks_numeric(42) is True

    def test_float(self):
        assert _looks_numeric(3.14) is True

    def test_string_number(self):
        assert _looks_numeric("42") is True

    def test_string_text(self):
        assert _looks_numeric("hello") is False

    def test_none(self):
        assert _looks_numeric(None) is False

    def test_empty_string(self):
        assert _looks_numeric("") is False


# ---------------------------------------------------------------------------
# Full Profile — Basic XLSX
# ---------------------------------------------------------------------------

class TestProfileXLSX:

    def test_basic_xlsx(self):
        rows = [
            ["id", "name", "score"],
            [1, "Alice", 95.5],
            [2, "Bob", 87.0],
            [3, "Charlie", 92.3],
        ]
        path = _create_xlsx(rows)
        try:
            raw_cols, row_count, is_exact = profile(path, SizeStrategy.MEMORY_SAFE)
            assert row_count == 3
            assert is_exact is True
            col_names = {c.name for c in raw_cols}
            assert "id" in col_names
            assert "name" in col_names
            assert "score" in col_names
        finally:
            os.unlink(path)

    def test_null_values(self):
        rows = [
            ["id", "name", "email"],
            [1, "Alice", "alice@test.com"],
            [2, "Bob", None],
            [3, None, None],
        ]
        path = _create_xlsx(rows)
        try:
            raw_cols, row_count, is_exact = profile(path, SizeStrategy.MEMORY_SAFE)
            assert row_count == 3

            email_col = next(c for c in raw_cols if c.name == "email")
            assert email_col.null_count == 2

            name_col = next(c for c in raw_cols if c.name == "name")
            assert name_col.null_count == 1
        finally:
            os.unlink(path)

    def test_boolean_values(self):
        rows = [
            ["id", "active"],
            [1, True],
            [2, False],
        ]
        path = _create_xlsx(rows)
        try:
            raw_cols, row_count, is_exact = profile(path, SizeStrategy.MEMORY_SAFE)
            active_col = next(c for c in raw_cols if c.name == "active")
            assert "true" in active_col.values
            assert "false" in active_col.values
        finally:
            os.unlink(path)

    def test_date_values(self):
        rows = [
            ["id", "created"],
            [1, datetime.datetime(2024, 1, 15, 10, 30, 0)],
            [2, datetime.datetime(2024, 6, 20, 14, 0, 0)],
        ]
        path = _create_xlsx(rows)
        try:
            raw_cols, row_count, is_exact = profile(path, SizeStrategy.MEMORY_SAFE)
            created_col = next(c for c in raw_cols if c.name == "created")
            assert any("2024-01-15" in v for v in created_col.values if v)
            assert any("2024-06-20" in v for v in created_col.values if v)
        finally:
            os.unlink(path)

    def test_float_integers_display_clean(self):
        """42.0 should be stored as '42', not '42.0'."""
        rows = [
            ["id", "count"],
            [1, 10.0],
            [2, 20.0],
        ]
        path = _create_xlsx(rows)
        try:
            raw_cols, row_count, is_exact = profile(path, SizeStrategy.MEMORY_SAFE)
            count_col = next(c for c in raw_cols if c.name == "count")
            assert "10" in count_col.values
            assert "20" in count_col.values
            assert "10.0" not in count_col.values
        finally:
            os.unlink(path)

    def test_empty_sheet(self):
        import openpyxl
        wb = openpyxl.Workbook()
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        # Active sheet exists but has no data
        wb.save(path)
        wb.close()
        path = Path(path)
        try:
            raw_cols, row_count, is_exact = profile(path, SizeStrategy.MEMORY_SAFE)
            assert row_count == 0
            assert raw_cols == []
        finally:
            os.unlink(path)

    def test_no_header_row(self):
        rows = [
            [1, 2, 3],
            [4, 5, 6],
            [7, 8, 9],
        ]
        path = _create_xlsx(rows)
        try:
            raw_cols, row_count, is_exact = profile(path, SizeStrategy.MEMORY_SAFE)
            assert row_count == 3  # all rows are data
            col_names = {c.name for c in raw_cols}
            assert "column_1" in col_names
            assert "column_2" in col_names
            assert "column_3" in col_names
        finally:
            os.unlink(path)


# ---------------------------------------------------------------------------
# Multi-sheet Files
# ---------------------------------------------------------------------------

class TestMultiSheet:

    def test_first_sheet_profiled(self):
        rows_sheet1 = [
            ["id", "name"],
            [1, "Alice"],
            [2, "Bob"],
        ]
        rows_sheet2 = [
            ["x", "y", "z"],
            [10, 20, 30],
        ]
        path = _create_xlsx(
            rows_sheet1,
            sheet_name="Main",
            extra_sheets={"Other": rows_sheet2},
        )
        try:
            raw_cols, row_count, is_exact = profile(path, SizeStrategy.MEMORY_SAFE)
            assert row_count == 2
            col_names = {c.name for c in raw_cols}
            assert "id" in col_names
            assert "name" in col_names
            # Should NOT contain columns from sheet 2
            assert "x" not in col_names
        finally:
            os.unlink(path)


# ---------------------------------------------------------------------------
# Sampling Strategies
# ---------------------------------------------------------------------------

class TestSamplingStrategies:

    def _make_large_xlsx(self, n: int = 200) -> Path:
        rows = [["id", "value"]]
        rows.extend([[i, f"val_{i}"] for i in range(n)])
        return _create_xlsx(rows)

    def test_memory_safe_reads_all(self):
        path = self._make_large_xlsx(50)
        try:
            raw_cols, row_count, is_exact = profile(path, SizeStrategy.MEMORY_SAFE)
            assert row_count == 50
            id_col = next(c for c in raw_cols if c.name == "id")
            assert len(id_col.values) == 50
        finally:
            os.unlink(path)

    def test_lazy_scan_uses_read_only(self):
        path = self._make_large_xlsx(50)
        try:
            raw_cols, row_count, is_exact = profile(path, SizeStrategy.LAZY_SCAN)
            assert row_count == 50
            # All 50 rows fit in reservoir (< SAMPLE_ROW_COUNT)
            id_col = next(c for c in raw_cols if c.name == "id")
            assert len(id_col.values) == 50
        finally:
            os.unlink(path)

    def test_stream_only_skip_interval(self):
        path = self._make_large_xlsx(200)
        try:
            raw_cols, row_count, is_exact = profile(path, SizeStrategy.STREAM_ONLY)
            assert row_count == 200
            id_col = next(c for c in raw_cols if c.name == "id")
            # Skip interval = 100 → should get rows at indices 0, 100
            assert len(id_col.values) == 2
        finally:
            os.unlink(path)


# ---------------------------------------------------------------------------
# Integration with main.profile_file
# ---------------------------------------------------------------------------

class TestIntegration:

    def test_profile_file_xlsx(self, tmp_path):
        from file_profiler.main import profile_file

        rows = [
            ["id", "name", "score"],
            [1, "Alice", 95],
            [2, "Bob", 87],
            [3, "Charlie", 92],
        ]
        xlsx_path = tmp_path / "test_data.xlsx"

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        wb.save(str(xlsx_path))
        wb.close()

        fp = profile_file(xlsx_path, output_dir=tmp_path)

        assert fp.file_format.value == "excel"
        assert fp.row_count == 3
        assert len(fp.columns) == 3
        col_names = {c.name for c in fp.columns}
        assert "id" in col_names
        assert "name" in col_names
        assert "score" in col_names

        # Verify output JSON was written
        output_file = tmp_path / "test_data_profile.json"
        assert output_file.exists()
